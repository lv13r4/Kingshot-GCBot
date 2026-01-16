#!/usr/bin/env python3
import tkinter as tk
from tkinter import messagebox
from tkinter.scrolledtext import ScrolledText
import requests
from bs4 import BeautifulSoup
import subprocess
import threading
import time
import os
import sys
from datetime import datetime
import queue
import discord
import asyncio
import openpyxl
import re
import json
import argparse

# -----------------------------
# ARGUMENT PARSING
# -----------------------------
parser = argparse.ArgumentParser(description="KS Gift Code Manager")
parser.add_argument("--headless", action="store_true", help="Run without GUI")
args = parser.parse_args()

# -----------------------------
# PATHS & CONFIGURATION
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")

def load_config():
    default_config = {
        "DISCORD_WEBHOOK": "",
        "MONITOR_TOKEN": "",
        "MONITOR_CHANNEL_ID": 0,
        "CHECK_INTERVAL": 3600,
        "TARGET_URL": "https://kingshot.net/gift-codes"
    }
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            try:
                user_config = json.load(f)
                return {**default_config, **user_config}
            except:
                return default_config
    return default_config

config = load_config()

DISCORD_WEBHOOK = config["DISCORD_WEBHOOK"]
MONITOR_TOKEN = config["MONITOR_TOKEN"]
MONITOR_CHANNEL_ID = config["MONITOR_CHANNEL_ID"]
CHECK_INTERVAL = config["CHECK_INTERVAL"]
TARGET_URL = config["TARGET_URL"]

REDEEM_SCRIPT = os.path.join(BASE_DIR, "redeemer.py")
CODES_FILE = os.path.join(BASE_DIR, "ks_codes.txt")
LOG_FILE = os.path.join(BASE_DIR, "ks_bot.log")
EXCEL_FILE = os.path.join(BASE_DIR, "KSGC.xlsx")
PYTHON_EXE = sys.executable

# -----------------------------
# THREAD-SAFE LOGGING
# -----------------------------
log_queue = queue.Queue()

def log(msg):
    timestamp = datetime.now().strftime("%H:%M:%S")
    line = f"[{timestamp}] {msg}\n"
    print(line, end="")
    log_queue.put(line)
    with open(LOG_FILE, "a") as f:
        f.write(line)

def process_log_queue():
    while not log_queue.empty():
        line = log_queue.get()
        if not args.headless:
            terminal_box.insert(tk.END, line)
            terminal_box.see(tk.END)
    if not args.headless:
        root.after(100, process_log_queue)

# -----------------------------
# DISCORD MONITOR CLIENT
# -----------------------------
class DiscordClient(discord.Client):
    def __init__(self, loop):
        intents = discord.Intents.default()
        intents.message_content = True
        super().__init__(intents=intents, loop=loop)

    async def on_ready(self):
        log(f"Discord Monitor: Logged in as {self.user} (ID: {self.user.id})")
        log(f"Discord Monitor: Listening to channel ID: {MONITOR_CHANNEL_ID}")

    async def on_message(self, message):
        if message.channel.id != MONITOR_CHANNEL_ID:
            return
        
        author = message.author.name
        content = message.content
        log(f"[CHAT] {author}: {content}")

        if content.strip().lower() == "check for codes":
            log(f"Received Check for codes command from {author}")
            try:
                await message.channel.send("Checking for new codes...")
                current_loop = asyncio.get_running_loop()
                def checker():
                    found_codes = check_new_codes()
                    if not found_codes:
                        asyncio.run_coroutine_threadsafe(
                            message.channel.send("No new codes found."), current_loop
                        )
                threading.Thread(target=checker, daemon=True).start()
            except Exception as e:
                log(f"Error triggering check: {e}")

        if content.strip().lower() == "redeem":
            log(f"Received Redeem command from {author}")
            try:
                await message.channel.send("Starting redeemer process...")
                current_loop = asyncio.get_running_loop()
                def runner():
                    if run_redeemer():
                        asyncio.run_coroutine_threadsafe(
                            message.channel.send("Redemption process completed."), current_loop
                        )
                    else:
                        asyncio.run_coroutine_threadsafe(
                            message.channel.send("Redeemer is already running or failed."), current_loop
                        )
                threading.Thread(target=runner, daemon=True).start()
            except Exception as e:
                log(f"Error triggering redeemer: {e}")

        match = re.match(r"Add Player ID:\s*(\S+)\s+(.*)", content, re.IGNORECASE)
        if match:
            player_id = match.group(1)
            player_name = match.group(2).strip()
            success = self.add_player_to_excel(player_id, player_name)
            if success:
                try:
                    await message.channel.send("Player ID added")
                except Exception as e:
                    log(f"Error sending Discord reply: {e}")

    def add_player_to_excel(self, player_id, player_name):
        try:
            if os.path.exists(EXCEL_FILE):
                workbook = openpyxl.load_workbook(EXCEL_FILE)
                sheet = workbook.active
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Players"

            for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                if row[0] and str(row[0]).strip() == player_id:
                    log(f"⚠️ Duplicate: Player ID {player_id} already exists.")
                    return False

            sheet.append([player_id, player_name])
            workbook.save(EXCEL_FILE)
            log(f"✅ Added Player: {player_name} (ID: {player_id})")
            return True
        except Exception as e:
            log(f"❌ Excel Error: {e}")
            return False

def start_discord_monitor():
    if not MONITOR_TOKEN:
        log("Discord Monitor: No token provided in config.json. Monitor disabled.")
        return
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    client = DiscordClient(loop)
    try:
        loop.run_until_complete(client.start(MONITOR_TOKEN))
    except Exception as e:
        log(f"Discord Monitor Error: {e}")

# -----------------------------
# CODE STORAGE
# -----------------------------
def load_sent_codes():
    if not os.path.exists(CODES_FILE):
        return set()
    with open(CODES_FILE, "r") as f:
        return set(line.strip() for line in f.readlines())

def save_sent_code(code):
    if os.path.exists(CODES_FILE) and os.path.getsize(CODES_FILE) > 0:
        with open(CODES_FILE, 'a+') as f:
            f.seek(f.tell() - 1, os.SEEK_SET)
            if f.read(1) != '\n':
                f.write('\n')
            f.write(code + "\n")
    else:
        with open(CODES_FILE, 'a') as f:
            f.write(code + "\n")

# -----------------------------
# DISCORD WEBHOOK
# -----------------------------
def send_discord_message(message):
    if not DISCORD_WEBHOOK:
        return
    data = {"content": message}
    try:
        requests.post(DISCORD_WEBHOOK, json=data, timeout=10)
    except Exception as e:
        log(f"Discord error: {e}")

# -----------------------------
# SCRAPER
# -----------------------------
def scrape_codes():
    try:
        r = requests.get(TARGET_URL, timeout=10)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        text_lines = soup.get_text(separator="\n").split("\n")

        codes = []
        in_active = False
        for line in text_lines:
            line = line.strip()
            if line.lower() == "active gift codes":
                in_active = True
                continue
            if line.lower() == "expired gift codes":
                in_active = False
            if in_active and line.isalnum() and 4 <= len(line) <= 20 and line.lower() != "active":
                codes.append(line)
        return list(dict.fromkeys(codes))
    except Exception as e:
        log(f"Scrape error: {e}")
        return []

# -----------------------------
# CHECK NEW CODES
# -----------------------------
def update_status(text):
    if not args.headless:
        root.after(0, lambda: status_label.config(text=f"Status: {text}"))

def check_new_codes():
    update_status("Checking...")
    sent_codes = load_sent_codes()
    scraped_codes = scrape_codes()
    added_codes = []

    for code in scraped_codes:
        if code not in sent_codes:
            send_discord_message(f"New Gift Code Found: `{code}`")
            save_sent_code(code)
            sent_codes.add(code)
            added_codes.append(code)
            log(f"New code found: {code}")

    if added_codes:
        log("New codes detected, starting redeemer automatically.")
        threading.Thread(target=run_redeemer, args=(added_codes,), daemon=True).start()
    else:
        log("No new codes found.")
    
    if not args.headless:
        root.after(0, lambda: last_check_label.config(
            text=f"Last Check: {datetime.now().strftime('%H:%M:%S')}"
        ))
    update_status("Idle")
    return added_codes

# -----------------------------
# RUN PYTHON REDEEMER
# -----------------------------
redeemer_lock = threading.Lock()

def run_redeemer(new_codes=None):
    if not redeemer_lock.acquire(blocking=False):
        log("Redeemer is already running.")
        return False

    try:
        update_status("Redeeming...")
        log("Running redeemer...")

        command = [PYTHON_EXE, REDEEM_SCRIPT, DISCORD_WEBHOOK]
        if new_codes:
            command.extend(new_codes)

        proc = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True
        )
        for line in proc.stdout:
            log(line.strip())
        proc.wait()
        
        if not args.headless:
            root.after(0, lambda: last_redeem_label.config(
                text=f"Last Redeem: {datetime.now().strftime('%H:%M:%S')}"
            ))
        return True
    except Exception as e:
        log(f"Redeemer error: {e}")
        return False
    finally:
        update_status("Idle")
        redeemer_lock.release()

# -----------------------------
# AUTO-CHECK LOOP
# -----------------------------
running = False

def start_auto_check():
    global running
    if not running:
        running = True
        threading.Thread(target=auto_check_loop, daemon=True).start()
        log("Auto-check started.")

def auto_check_loop():
    while running:
        check_new_codes()
        for _ in range(int(CHECK_INTERVAL)):
            if not running:
                break
            time.sleep(1)

# -----------------------------
# MAIN EXECUTION
# -----------------------------
if args.headless:
    log("Running in HEADLESS mode (No GUI).")
    start_auto_check()
    threading.Thread(target=start_discord_monitor, daemon=True).start()
    try:
        while True:
            time.sleep(1)
            process_log_queue()
    except KeyboardInterrupt:
        log("Stopping bot...")
        running = False
else:
    root = tk.Tk()
    root.title("KS Gift Code Manager")
    root.geometry("600x400")

    status_label = tk.Label(root, text="Status: Idle", anchor="w")
    status_label.pack(fill="x", padx=10, pady=5)

    last_check_label = tk.Label(root, text="Last Check: Never", anchor="w")
    last_check_label.pack(fill="x", padx=10, pady=5)

    last_redeem_label = tk.Label(root, text="Last Redeem: Never", anchor="w")
    last_redeem_label.pack(fill="x", padx=10, pady=5)

    button_frame = tk.Frame(root)
    button_frame.pack(padx=10, pady=5)

    tk.Button(button_frame, text="Check for New Codes", width=20,
              command=lambda: threading.Thread(target=check_new_codes, daemon=True).start()).grid(row=0, column=0, padx=5, pady=5)

    tk.Button(button_frame, text="Redeem Codes", width=20,
              command=lambda: threading.Thread(target=run_redeemer, daemon=True).start()).grid(row=0, column=1, padx=5, pady=5)

    terminal_box = ScrolledText(root, height=12, bg="black", fg="lime", insertbackground="lime")
    terminal_box.pack(fill="both", expand=True, padx=10, pady=5)
    terminal_box.insert(tk.END, "=== KS Bot Log ===\n")

    def on_close():
        global running
        running = False
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    process_log_queue()
    threading.Thread(target=start_discord_monitor, daemon=True).start()
    start_auto_check()
    root.mainloop()
